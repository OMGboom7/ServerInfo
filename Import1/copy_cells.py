import openpyxl
from openpyxl.reader.excel import load_workbook


def copy_cell(excel_file_path):
    wb = load_workbook(excel_file_path)
    ws = wb.active

    start_cell = 'A3'
    end_cell = 'L5'
    cell_range = ws.cell()



if __name__ == '__main__':
    copy_cell()
