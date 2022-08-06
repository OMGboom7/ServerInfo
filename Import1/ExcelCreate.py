import os

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment, Side

from Import1 import GetPassword


def create_excel():
    wb = openpyxl.Workbook()
    ws = wb.active

    font_tile = Font(
        name="微软雅黑",
        size=20,
        color="000000"
    )
    font_normal = Font(
        name="微软雅黑",
        size=10,
        color="000000"
    )
    alignment_title = Alignment(
        vertical='center',
        horizontal='centerContinuous'
    )
    side_sheet = Side(
        style='thin',
        color='000000'
    )
    border_sheet = Border(
        left=side_sheet,
        top=side_sheet,
        right=side_sheet,
        bottom=side_sheet
    )
    patternfill_title = PatternFill(
        fill_type='solid',
        fgColor='80bfff'
    )

    # 工作表标题
    ws.title = '清单'

    ws.merge_cells('A1:L2')
    ws.merge_cells('B3:C3')
    ws.merge_cells('B4:C4')
    ws.merge_cells('B5:C5')
    ws.merge_cells('E3:F3')
    ws.merge_cells('E4:F4')
    ws.merge_cells('E5:F5')
    ws.merge_cells('H3:I3')
    ws.merge_cells('H4:I4')
    ws.merge_cells('H5:I5')
    ws.merge_cells('K3:L3')
    ws.merge_cells('K4:L4')
    ws.merge_cells('K5:L5')

    ws['A3'] = 'IP'
    ws['A4'] = '运行时间'
    ws['A5'] = '上行速率'
    ws['D3'] = '系统版本'
    ws['D4'] = '系统负载'
    ws['D5'] = '下行速率'
    ws['G3'] = '主机名'
    ws['G4'] = '内存容量'
    ws['G5'] = '存储容量'
    ws['J3'] = 'CPU使用率'
    ws['J4'] = '内存使用率'
    ws['J5'] = '存储使用率'

    # 列宽11
    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['D'].width = 11
    ws.column_dimensions['G'].width = 11
    ws.column_dimensions['J'].width = 11

    # 批量居中
    for rows1 in range(0, 100):
        for cols1 in range(0, 100):
            ws.cell(row=rows1 + 1, column=cols1 + 1).alignment = Alignment(horizontal='center', vertical='center')

    # 边框
    border1 = ws['A1:L5']
    for k in border1:
        for k1 in k:
            k1.border = border_sheet
            k1.font = font_normal

    info = GetPassword.get_yaml_data('config.yaml')
    server_num = len(info)
    ws['A1'] = '服务器资源清单（共' + str(server_num) + '台）'
    ws['A1'].font = font_tile
    ws['A1'].alignment = alignment_title
    # ws['A1'].border = border_sheet
    ws['A1'].fill = patternfill_title

    if not os.path.exists('../Excel'):
        os.mkdir('../Excel')

    wb.save("../Excel/服务器资源清单（共" + str(server_num) + "台）.xlsx")



if __name__ == "__main__":
    create_excel()
